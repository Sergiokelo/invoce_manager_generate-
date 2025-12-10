import os 
import re
import random
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- géométrie (centroïdes) ---
try:
    from shapely.wkt import loads as wkt_loads
    from shapely.ops import unary_union
    HAS_SHAPELY = True
except Exception:
    HAS_SHAPELY = False


# ---------- OUTILS ----------
def reduit(val: str) -> str:
    if isinstance(val, str) and len(val) >= 3:
        return val[:2].upper() + val[-1].upper()
    return "UNK"

def ask(path_default, prompt):
    txt = input(f"{prompt} (Entrée pour défaut)\n> ").strip()
    return txt or path_default

def yn(prompt, default=True):
    d = "[O/n]" if default else "[o/N]"
    ans = input(f"{prompt} {d} ").strip().lower()
    if not ans:
        return default
    return ans in ("o", "oui", "y", "yes")

def clean_path(p: str) -> str:
    if p is None:
        return p
    p = p.strip().strip("'\"")
    return os.path.normpath(p)

def ensure_dir(p: str) -> str:
    """Crée le dossier parent du chemin fichier, retourne le chemin nettoyé."""
    p = clean_path(p)
    os.makedirs(os.path.dirname(p), exist_ok=True)
    return p

def resolve_excel_path(path_like: str) -> str:
    """Accepte un chemin fichier .xls(x/m) ou un dossier.
       Si dossier: liste les fichiers Excel et laisse l'utilisateur choisir."""
    p = clean_path(path_like)
    if os.path.isdir(p):
        cands = [os.path.join(p, f) for f in os.listdir(p)
                 if f.lower().endswith(('.xlsx', '.xls', '.xlsm')) and not f.startswith('~$')]
        if not cands:
            raise FileNotFoundError(f"Aucun fichier Excel trouvé dans : {p}")
        if len(cands) == 1:
            print(f"→ Fichier détecté : {cands[0]}")
            return cands[0]
        print("Plusieurs fichiers trouvés :")
        for i, f in enumerate(cands, 1):
            print(f"  {i}. {os.path.basename(f)}")
        while True:
            sel = input("Choisis le numéro du fichier à utiliser : ").strip()
            if sel.isdigit() and 1 <= int(sel) <= len(cands):
                return cands[int(sel) - 1]
            print("Numéro invalide, réessaie.")
    return p


# ---------- INTERACTIF ----------
print("\n=== Générateur interactif Files / AgriTrace ===\n")

# chemins par défaut (modifiables à l'invite)
input_file_path_raw = ask(r"D:\doc\traçabilité\MCK COOP CA\phase 2 mapping\TABLE D'ATTRIBUT MCK 2EME PHASE.xlsx",
                          "Chemin du fichier source Excel")
sheet_name = ask("Feuil1", "Nom de la feuille Excel (sheet_name)")
out_dir_raw = ask(r"D:\doc\traçabilité\MCK COOP CA\phase 2 mapping", "Dossier de sortie")

# nettoie / résout
input_file_path = resolve_excel_path(input_file_path_raw)
out_dir = clean_path(out_dir_raw)

print(f"→ Fichier source: {input_file_path}")
print(f"→ Dossier sortie: {out_dir}")

cleaned_file_path   = ensure_dir(os.path.join(out_dir, "TABLE -P1.xlsx"))
upload_file_path    = ensure_dir(os.path.join(out_dir, "TABLE MAP UPLOAD.xlsx"))
result_path         = ensure_dir(os.path.join(out_dir, "base de donnee.xlsx"))
multi_zone_path     = ensure_dir(os.path.join(out_dir, "PLANTEUR_MULTI_ZONE.xlsx"))
minimal_extra_path  = ensure_dir(os.path.join(out_dir, "PLANTEUR_MINIMAL_CENTROID.xlsx"))

gen_upload   = yn("Générer le fichier d’upload (avec CODE_PLANTATION, RENDEMENT, etc.) ?", True)
gen_summary  = yn("Générer le résumé planteurs + MULTI_ZONE (avec coloration) ?", True)
gen_minimal  = yn("Générer le fichier supplémentaire (colonnes demandées + centroïdes) ?", True)

print("\nLecture du fichier source…")
# lecture Excel avec fallback si le nom de feuille n'existe pas
try:
    df = pd.read_excel(input_file_path, sheet_name=sheet_name)
except ValueError:
    xls = pd.ExcelFile(input_file_path)
    print(f'⚠️ Feuille "{sheet_name}" introuvable. Utilisation de : {xls.sheet_names[0]}')
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0])

# ---------- 1) Nettoyage initial ----------
print("Étape 1/4 : Nettoyage et transformation du CODE…")
# enlever le suffixe -P1, -P2, ...
df['CODE'] = df['CODE'].astype(str).str.replace(r'-P\d+', '', regex=True)

# NUMERO_PLANTEUR (numérique à la fin du CODE)
df['NUMERO_PLANTEUR'] = df['CODE'].str.extract(r'-(\d+)$')
# COOP = préfixe avant le 1er '-'
df['COOP'] = df['CODE'].str.extract(r'^([^-]+)')

# Réductions AXE / VILLAGE
if 'AXE' not in df.columns:
    df['AXE'] = "UNK"
if 'VILLAGE' not in df.columns:
    df['VILLAGE'] = "UNK"

df['AXE_REDUIT'] = df['AXE'].apply(reduit)
df['VILLAGE_REDUIT'] = df['VILLAGE'].apply(reduit)

# Nouveau CODE unifié (sans -P#)
df['CODE'] = df['COOP'] + '-' + df['AXE_REDUIT'] + '-' + df['VILLAGE_REDUIT'] + '-' + df['NUMERO_PLANTEUR']

df.to_excel(cleaned_file_path, index=False)
print(f"✅ Nettoyage terminé → {cleaned_file_path}")

# ---------- 2) Fichier UPLOAD ----------
if gen_upload:
    print("Étape 2/4 : Génération du fichier d’upload…")
    df_u = pd.read_excel(cleaned_file_path)

    # numérotation des plantations P1, P2, ... au niveau du NUMERO_PLANTEUR
    df_u['Occurrence'] = df_u.groupby('NUMERO_PLANTEUR').cumcount() + 1
    df_u['CODE_PLANTATION'] = df_u['CODE'] + '-P' + df_u['Occurrence'].astype(str)

    # colonnes utiles pour le rendement
    if 'HECTARE' not in df_u.columns and 'SUPERFICIE' in df_u.columns:
        df_u['HECTARE'] = df_u['SUPERFICIE']

    df_u['RENDEMENT'] = df_u.get('HECTARE', pd.Series([0]*len(df_u))) * 950
    df_u['CERTIFICATION'] = "NONE"

    df_u = df_u.drop(columns=[c for c in ['Occurrence', 'COOP', 'AXE_REDUIT', 'VILLAGE_REDUIT'] if c in df_u.columns])

    # réorganiser : CODE_PLANTATION juste après CODE
    cols = list(df_u.columns)
    if 'CODE_PLANTATION' in cols and 'CODE' in cols:
        cols.insert(cols.index('CODE') + 1, cols.pop(cols.index('CODE_PLANTATION')))
        df_u = df_u[cols]

    df_u.to_excel(upload_file_path, index=False)
    print(f"✅ Fichier d’upload généré → {upload_file_path}")

# ---------- 3) Résumé planteur + MULTI_ZONE ----------
planteurs_multi_zone = []
resultat = None

if gen_summary or gen_minimal:
    print("Étape 3/4 : Calculs agrégés par planteur…")
    df_c = pd.read_excel(cleaned_file_path)

    # agrégats par CODE (clé planteur)
    planteur_key = 'CODE'

    # fallback HECTARE si absent
    if 'HECTARE' not in df_c.columns and 'SUPERFICIE' in df_c.columns:
        df_c['HECTARE'] = df_c['SUPERFICIE']

    # Nombre de plantation par CODE
    df_c['Nombre de plantation'] = df_c.groupby(planteur_key)[planteur_key].transform('count')

    # Somme des surfaces par CODE
    for col in ['SUPERFICIE', 'HECTARE']:
        if col in df_c.columns:
            df_c[col] = df_c.groupby(planteur_key)[col].transform('sum')

    # dédupliquer : 1 ligne par CODE planteur
    resultat = df_c.drop_duplicates(subset=planteur_key).reset_index(drop=True)

    # ANNEE_NAISSANCE + GENRE si absents
    if 'ANNEE_NAISSANCE' not in resultat.columns:
        resultat['ANNEE_NAISSANCE'] = [datetime.now().year - random.randint(20, 50) for _ in range(len(resultat))]
    if 'GENRE' not in resultat.columns:
        resultat['GENRE'] = [random.choices(['M', 'F'], weights=[60, 40])[0] for _ in range(len(resultat))]

    # MULTI_ZONE (variation AXE/VILLAGE au niveau NUMERO_PLANTEUR)
    df_check = pd.read_excel(cleaned_file_path)
    for col in ['AXE', 'VILLAGE']:
        if col not in df_check.columns:
            df_check[col] = "UNK"
    if 'NUMERO_PLANTEUR' in df_check.columns:
        df_check['NUMERO_PLANTEUR'] = df_check['NUMERO_PLANTEUR'].astype(str)

    variation_detectee = (
        df_check.groupby('NUMERO_PLANTEUR')[['AXE', 'VILLAGE']]
        .nunique()
        .apply(lambda r: (r['AXE'] > 1) or (r['VILLAGE'] > 1), axis=1)
    )
    # liste des NUMERO_PLANTEUR multi-zone (en str)
    planteurs_multi_zone = variation_detectee[variation_detectee].index.astype(str).tolist()

    # s'assurer que NUMERO_PLANTEUR est bien présent et en str dans le résumé
    if 'NUMERO_PLANTEUR' in resultat.columns:
        resultat['NUMERO_PLANTEUR'] = resultat['NUMERO_PLANTEUR'].astype(str)
    else:
        # on le recalcule à partir de CODE si besoin
        resultat['NUMERO_PLANTEUR'] = resultat['CODE'].str.extract(r'-(\d+)$')[0].astype(str)

    resultat['MULTI_ZONE'] = resultat['NUMERO_PLANTEUR'].apply(
        lambda x: "OUI" if x in planteurs_multi_zone else "NON"
    )

    # ordre de colonnes : 1 ligne par CODE planteur
    cols_order = ['CODE', 'NUMERO_PLANTEUR', 'NOM', 'GENRE', 'ANNEE_NAISSANCE',
                  'Nombre de plantation', 'SUPERFICIE', 'HECTARE', 'MULTI_ZONE']
    resultat = resultat[[c for c in cols_order if c in resultat.columns]].copy()

    # harmoniser types pour merges
    if 'CODE' in resultat.columns:
        resultat['CODE'] = resultat['CODE'].astype(str)
    if 'NUMERO_PLANTEUR' in resultat.columns:
        resultat['NUMERO_PLANTEUR'] = resultat['NUMERO_PLANTEUR'].astype(str)

    if gen_summary:
        resultat.to_excel(result_path, index=False)

        # coloration verte pour multi-zone (sur base de la colonne MULTI_ZONE)
        wb = load_workbook(result_path)
        ws = wb.active
        fill_vert = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # repérer l'index de la colonne MULTI_ZONE
        header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        header_names = [cell.value for cell in header_row]
        try:
            idx_multi = header_names.index('MULTI_ZONE')
        except ValueError:
            idx_multi = None

        if idx_multi is not None:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell_multi = row[idx_multi]
                if cell_multi.value == "OUI":
                    for cell in row:
                        cell.fill = fill_vert

        wb.save(result_path)
        print(f"✅ Résumé + coloration enregistrés → {result_path}")

        # fichier multi-zone seul
        if 'MULTI_ZONE' in resultat.columns:
            multi_zone_df = resultat[resultat['MULTI_ZONE'] == 'OUI']
            multi_zone_df.to_excel(multi_zone_path, index=False)
            print(f"✅ Fichier des multi-zone → {multi_zone_path}")

# ---------- 4) FICHIER SUPPLÉMENTAIRE (colonnes demandées + centroïdes) ----------
if gen_minimal:
    print("Étape 4/4 : Génération du fichier supplémentaire (colonnes demandées)…")

    if not HAS_SHAPELY:
        raise RuntimeError(
            "Shapely n’est pas installé. Installe-le avec: pip install shapely\n"
            "Ce module est requis pour calculer long/lat depuis wkt_geom."
        )

    base = pd.read_excel(cleaned_file_path)

    # trouver colonnes candidates
    col_nom   = 'NOM' if 'NOM' in base.columns else None
    col_sect  = next((c for c in ["Secteur d'activité", "SECTEUR_ACTIVITE", "SECTEUR", "SECTEUR D'ACTIVITE"]
                      if c in base.columns), None)
    col_crops = next((c for c in ["crops_grown", "CULTURE", "CULTURE_PRINCIPALE", "PRODUCTION"]
                      if c in base.columns), None)

    # centroïdes par CODE = union des géométries des parcelles de ce code planteur
    if 'wkt_geom' not in base.columns:
        base['wkt_geom'] = None

    def geom_or_none(w):
        try:
            return wkt_loads(w) if isinstance(w, str) and w.strip() else None
        except Exception:
            return None

    base['__geom'] = base['wkt_geom'].apply(geom_or_none)

    # harmoniser types de clé
    if 'CODE' in base.columns:
        base['CODE'] = base['CODE'].astype(str)

    centroids = []
    grouped = base.groupby('CODE')
    for code_pl, sub in grouped:
        geoms = [g for g in sub['__geom'].tolist() if g is not None]
        if geoms:
            try:
                union = unary_union(geoms) if len(geoms) > 1 else geoms[0]
                c = union.centroid
                centroids.append((code_pl, c.x, c.y))
            except Exception:
                centroids.append((code_pl, None, None))
        else:
            centroids.append((code_pl, None, None))

    cent_df = pd.DataFrame(centroids, columns=['CODE_PLANTEUR', 'long', 'lat'])

    # ✅ forcer numérique et ARRONDIR long/lat à 5 décimales dès la source
    cent_df['long'] = pd.to_numeric(cent_df['long'], errors='coerce').round(5)
    cent_df['lat']  = pd.to_numeric(cent_df['lat'],  errors='coerce').round(5)

    # base info par planteur : 1 ligne par CODE
    first_by = base.sort_values('CODE').drop_duplicates('CODE')

    minimal = first_by[['CODE']].copy()
    minimal.rename(columns={'CODE': 'CODE_PLANTEUR'}, inplace=True)

    # NOM
    minimal['NOM'] = first_by[col_nom] if col_nom else "ND"

    # Genre + ANNEE_NAISSANCE depuis 'resultat' si dispo (clé = CODE)
    if (resultat is not None) and ('CODE' in resultat.columns):
        cols_to_merge = [c for c in ['CODE', 'GENRE', 'ANNEE_NAISSANCE'] if c in resultat.columns]
        tmp = resultat[cols_to_merge].rename(columns={'CODE': 'CODE_PLANTEUR'})
        minimal = minimal.merge(tmp, on='CODE_PLANTEUR', how='left')

    # fallback si GENRE / ANNEE_NAISSANCE manquants
    if 'GENRE' not in minimal.columns:
        minimal['GENRE'] = first_by['GENRE'] if 'GENRE' in first_by.columns else "ND"
    if 'ANNEE_NAISSANCE' not in minimal.columns:
        minimal['ANNEE_NAISSANCE'] = [datetime.now().year - random.randint(20, 50) for _ in range(len(minimal))]

    # ✅ experience_years = ANNEE_NAISSANCE (année de naissance)
    minimal['experience_years'] = pd.to_numeric(minimal['ANNEE_NAISSANCE'], errors='coerce').astype('Int64')

    # Secteur d'activité
    minimal["Secteur d'activité"] = first_by[col_sect] if col_sect else "Agriculture"

    # crops_grown
    minimal['crops_grown'] = first_by[col_crops] if col_crops else "cacao"

    # coords depuis centroïdes (clé = CODE_PLANTEUR)
    minimal = minimal.merge(cent_df, on='CODE_PLANTEUR', how='left')

    # colonnes finales
    final_cols = ['CODE_PLANTEUR', 'NOM', 'GENRE', "Secteur d'activité",
                  'crops_grown', 'experience_years', 'long', 'lat']
    for col in final_cols:
        if col not in minimal.columns:
            minimal[col] = "" if col not in ('experience_years', 'long', 'lat') else None

    minimal = minimal[final_cols]

    # ✅ sécurité : re-arrondir après le merge, avant export
    minimal['long'] = pd.to_numeric(minimal['long'], errors='coerce').round(5)
    minimal['lat']  = pd.to_numeric(minimal['lat'],  errors='coerce').round(5)

    minimal.to_excel(minimal_extra_path, index=False)
    print(f"✅ Fichier supplémentaire généré → {minimal_extra_path}")

    # ✅ appliquer un format "0.00000" dans Excel pour long/lat (affichage)
    wb = load_workbook(minimal_extra_path)
    ws = wb.active

    # Index 1-based des colonnes en lisant l'en-tête
    header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
    col_index = {cell.value: (idx + 1) for idx, cell in enumerate(header_row)}

    for name in ('long', 'lat'):
        idx = col_index.get(name)
        if idx:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    cell.number_format = "0.00000"

    wb.save(minimal_extra_path)

print("\n🎉 Terminé.")
