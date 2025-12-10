import pandas as pd
import re
import random
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === 1. Nettoyage initial du fichier source ===
input_file_path = r"D:\doc\traçabilité\COPROCAB\REDIT_TEST\TABLE D'ATTRIBUT COOPROCABE 2EME PHASE.xlsx"
cleaned_file_path = r"D:\doc\traçabilité\COPROCAB\REDIT_TEST\Table-P1.xlsx"

df = pd.read_excel(input_file_path, sheet_name='Feuil1')

# Supprimer les suffixes "-P1", "-P2", etc.
df['CODE'] = df['CODE'].str.replace(r'-P\d+', '', regex=True)

# Extraire NUMERO_PLANTEUR
df['NUMERO_PLANTEUR'] = df['CODE'].str.extract(r'-(\d+)$')

# Extraire la coopérative
df['COOP'] = df['CODE'].str.extract(r'^([^-]+)')

# Réductions pour AXE et VILLAGE
def reduit(val):
    if isinstance(val, str) and len(val) >= 3:
        return val[:2].upper() + val[-1].upper()
    return "UNK"

df['AXE_REDUIT'] = df['AXE'].apply(reduit)
df['VILLAGE_REDUIT'] = df['VILLAGE'].apply(reduit)

# Nouveau CODE
df['CODE'] = df['COOP'] + '-' + df['AXE_REDUIT'] + '-' + df['VILLAGE_REDUIT'] + '-' + df['NUMERO_PLANTEUR']

# Sauvegarder fichier nettoyé
df.to_excel(cleaned_file_path, index=False)
print("✅ Étape 1 : Nettoyage et transformation du CODE terminé")

# === 2. Ajout CODE_PLANTATION, RENDEMENT, CERTIFICATION ===
df = pd.read_excel(cleaned_file_path)

df['Occurrence'] = df.groupby('NUMERO_PLANTEUR').cumcount() + 1
df['CODE_PLANTATION'] = df['CODE'] + '-P' + df['Occurrence'].astype(str)
df['RENDEMENT'] = df['HECTARE'] * 950
df['CERTIFICATION'] = "NONE"

# Supprimer colonnes inutiles
df = df.drop(columns=['Occurrence', 'COOP', 'AXE_REDUIT', 'VILLAGE_REDUIT'])

# Réorganiser
cols = list(df.columns)
cols.insert(cols.index('CODE') + 1, cols.pop(cols.index('CODE_PLANTATION')))
df = df[cols]

upload_file_path = r"D:\doc\traçabilité\COPROCAB\REDIT_TEST\TABLE MAP UPLOAD.xlsx"
df.to_excel(upload_file_path, index=False)
print("✅ Étape 2 : Fichier upload généré avec succès")

# === 3. PLANTEUR_X_PLANTATION + Multi-zone ===
df = pd.read_excel(cleaned_file_path)

# Agrégation
df['Nombre de plantation'] = df.groupby('NUMERO_PLANTEUR')['NUMERO_PLANTEUR'].transform('count')
df['SUPERFICIE'] = df.groupby('NUMERO_PLANTEUR')['SUPERFICIE'].transform('sum')
df['HECTARE'] = df.groupby('NUMERO_PLANTEUR')['HECTARE'].transform('sum')

# Supprimer doublons par planteur
resultat = df.drop_duplicates(subset='NUMERO_PLANTEUR').reset_index(drop=True)

# Ajout GENRE et ANNEE_NAISSANCE
resultat['ANNEE_NAISSANCE'] = [datetime.now().year - random.randint(20, 50) for _ in range(len(resultat))]
resultat['GENRE'] = [random.choices(['M', 'F'], weights=[60, 40])[0] for _ in range(len(resultat))]

# Identifier MULTI_ZONE
df_check = pd.read_excel(cleaned_file_path)
variation_detectee = (
    df_check.groupby('NUMERO_PLANTEUR')[['AXE', 'VILLAGE']]
    .nunique()
    .apply(lambda row: row['AXE'] > 1 or row['VILLAGE'] > 1, axis=1)
)
planteurs_multi_zone = variation_detectee[variation_detectee].index.tolist()
resultat['MULTI_ZONE'] = resultat['NUMERO_PLANTEUR'].apply(lambda x: "OUI" if x in planteurs_multi_zone else "NON")

# Réorganisation colonnes
cols_order = ['CODE', 'NOM', 'GENRE', 'ANNEE_NAISSANCE',
              'Nombre de plantation', 'SUPERFICIE', 'HECTARE', 'MULTI_ZONE']
resultat = resultat[cols_order]

# Sauvegarde fichier principal
result_path = r"D:\doc\traçabilité\COPROCAB\REDIT_TEST\base de donnee.xlsx"
resultat.to_excel(result_path, index=False)

# Générer fichier des planteurs multi-zone uniquement
multi_zone_df = resultat[resultat['MULTI_ZONE'] == 'OUI']
multi_zone_path = r"D:\doc\traçabilité\COPROCAB\REDIT_TEST\PLANTEUR_MULTI_ZONE.xlsx"
multi_zone_df.to_excel(multi_zone_path, index=False)

# === 4. Coloration verte dans le fichier final ===
wb = load_workbook(result_path)
ws = wb.active
fill_vert = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    code = row[0].value
    numero = code.split('-')[-1] if code else None
    if numero in planteurs_multi_zone:
        for cell in row:
            cell.fill = fill_vert

wb.save(result_path)

print("✅ Étape 3 : Résumé planteurs + MULTI_ZONE ajouté + coloration verte appliquée")
print("📄 Fichier final :", result_path)
print("📄 Fichier des multi-zone :", multi_zone_path)
print("\n🎉 Tous les fichiers ont été générés avec succès.")
