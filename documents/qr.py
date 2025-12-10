import os
import qrcode
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Chemin absolu vers le fichier source
input_path = r"D:\doc\traçabilité\SYCOODEP\DATA\TABLE6.xlsx"
output_path = r"D:\doc\traçabilité\SYCOODEP\DATA\TABLE6_with_QR2.xlsx"

temp_dir = r"D:/temp/"

# Créer le dossier temporaire si nécessaire
os.makedirs(temp_dir, exist_ok=True)

# Charger le fichier Excel
wb = load_workbook(input_path)
ws = wb.active  # Utiliser la feuille active

# Ajouter une colonne pour les QR codes
qr_column = ws.max_column + 1
ws.cell(row=1, column=qr_column, value="QR_Code")  # Titre de la colonne

# Dimensions souhaitées pour les QR codes (en pixels)
qr_width = 80
qr_height = 80

# Générer et insérer des QR codes
for row in range(2, ws.max_row + 1):
    # Récupérer les données de la ligne
    code = ws.cell(row=row, column=1).value  # Colonne A
    nom = ws.cell(row=row, column=2).value  # Colonne B
    plantations = ws.cell(row=row, column=3).value  # Colonne C
    
    # Contenu du QR code
    content = f"CODE: {code}, NOM: {nom}, Nombre de plantation: {plantations}"
    
    # Générer le QR code
    qr = qrcode.QRCode(box_size=10, border=4)
    qr.add_data(content)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Sauvegarder le QR code en tant qu'image temporaire
    img_path = f"{temp_dir}qr_code_{row}.png"
    img.save(img_path)
    
    # Insérer et redimensionner l'image dans la cellule Excel
    qr_image = Image(img_path)
    qr_image.width = qr_width  # Largeur redimensionnée
    qr_image.height = qr_height  # Hauteur redimensionnée
    
    # Déterminer la position de la cellule pour insérer l'image
    cell_position = f"{chr(64 + qr_column)}{row}"  # Convertir la colonne en lettre (par exemple, D2)
    ws.add_image(qr_image, cell_position)

# Sauvegarder le fichier Excel avec QR codes redimensionnés
wb.save(output_path)

print(f"Fichier sauvegardé avec QR codes redimensionnés : {output_path}")
