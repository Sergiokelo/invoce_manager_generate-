
import io, os, sqlite3, calendar, random
from datetime import date, datetime
from pathlib import Path
import numpy as np
import pandas as pd
import streamlit as st

# -------------------- Page --------------------
st.set_page_config(page_title="Planificateur des livraisons (rendement annuel)", layout="wide")
st.title("📦 Planificateur des livraisons par plantation — version avancée")
st.caption("Saisonnalité 60/40, Target, fréquences aléatoires, ordre intercalé, calendrier & cap, tolérance, exclusions, historique SQLite, fiche d’export.")

DB_PATH = str(Path(__file__).resolve().parent / "state.sqlite")

# -------------------- Base SQLite --------------------
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
        CREATE TABLE IF NOT EXISTS livraisons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code_plantation TEXT NOT NULL,
            nom_plantation TEXT,
            cooperative TEXT,
            annee INTEGER NOT NULL,
            mois INTEGER NOT NULL,
            date_livraison TEXT NOT NULL,
            quantite_kg REAL NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        c.execute("CREATE INDEX IF NOT EXISTS idx_livr_ym_code ON livraisons(annee, mois, code_plantation)")
        conn.commit()

def exists_delivery(code, annee, mois, dte, qty):
    with sqlite3.connect(DB_PATH) as conn:
        df = pd.read_sql_query(
            "SELECT COUNT(*) n FROM livraisons WHERE code_plantation=? AND annee=? AND mois=? AND date_livraison=? AND quantite_kg=?",
            conn, params=(str(code), int(annee), int(mois), str(dte), float(qty)))
    return int(df.iloc[0]["n"]) > 0

def insert_delivery(code, nom, coop, annee, mois, dte, qty):
    if exists_delivery(code, annee, mois, dte, qty):
        return False
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""INSERT INTO livraisons (code_plantation, nom_plantation, cooperative, annee, mois, date_livraison, quantite_kg)
                        VALUES (?,?,?,?,?,?,?)""", (str(code), nom, coop, int(annee), int(mois), str(dte), float(qty)))
        conn.commit()
    return True

def load_delivered_by_month(annee, mois):
    with sqlite3.connect(DB_PATH) as conn:
        q = "SELECT code_plantation, SUM(quantite_kg) sumq FROM livraisons WHERE annee=? AND mois=? GROUP BY code_plantation"
        df = pd.read_sql_query(q, conn, params=(int(annee), int(mois)))
    return dict(zip(df["code_plantation"], df["sumq"])) if len(df) else {}

def load_delivered_by_year(annee):
    with sqlite3.connect(DB_PATH) as conn:
        q = "SELECT code_plantation, SUM(quantite_kg) sumq FROM livraisons WHERE annee=? GROUP BY code_plantation"
        df = pd.read_sql_query(q, conn, params=(int(annee),))
    return dict(zip(df["code_plantation"], df["sumq"])) if len(df) else {}

def fetch_history(limit=200000):
    with sqlite3.connect(DB_PATH) as conn:
        df = pd.read_sql_query("SELECT * FROM livraisons ORDER BY created_at DESC LIMIT ?", conn, params=(int(limit),))
    return df

def clear_history():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("DELETE FROM livraisons")
        conn.commit()

init_db()

# -------------------- UI (barre latérale) --------------------
st.sidebar.header("1) Fichier source plantations")
st.sidebar.write("**Excel requis :** `CODE_PLANTATION`, `CODE`, `RENDEMENT`  — `NOM` **optionnel** (si absent ⇒ `NOM = CODE`)  — `COOPÉRATIVE` optionnel.")
uploaded = st.sidebar.file_uploader("Glisser un .xlsx (200 MB max)", type=["xlsx"])

st.sidebar.header("2) Période & fréquence")
annee = st.sidebar.number_input("Année", min_value=2020, max_value=2035, value=2025, step=1)
mois_label = st.sidebar.selectbox("Mois précis",
                                  ["(Toute l'année)","Janvier","Février","Mars","Avril","Mai","Juin",
                                   "Juillet","Août","Septembre","Octobre","Novembre","Décembre"], index=0)
freq_global = st.sidebar.slider("Fréquence globale (fallback)", min_value=1, max_value=5, value=2, step=1)

st.sidebar.header("3) Fréquence aléatoire par plantation")
freq_rand = st.sidebar.checkbox("Activer fréquence aléatoire (par plantation)", value=True)
minmax = st.sidebar.slider("Plage de fréquence (livraisons / mois)", 1, 5, (1, 3), disabled=not freq_rand)
freq_min, freq_max = minmax if freq_rand else (freq_global, freq_global)

st.sidebar.header("4) Target / Traçage global (optionnel)")
trace_global = st.sidebar.checkbox("Activer 'Quantité à tracer' (kg)")
quota_global = st.sidebar.number_input("Quantité à tracer (kg)", min_value=0.0, value=24000.0, step=100.0, disabled=not trace_global)
mode = st.sidebar.selectbox("Mode répartition Target", ["Proportionnel (par défaut)", "Sous-ensemble minimal (greedy)", "Échantillonnage pondéré"])
seuil_min = st.sidebar.number_input("Seuil min (kg) par plantation", min_value=0.0, value=0.0, step=50.0)

st.sidebar.header("5) Contraintes calendrier")
avoid_weekend = st.sidebar.checkbox("Éviter les week-ends (samedi/dimanche)", value=True)
min_gap_days = st.sidebar.number_input("Écart minimum entre deux livraisons d'une même plantation (jours)", min_value=0, max_value=15, value=3, step=1)
st.sidebar.caption("Ex. 3 jours: si une livraison est le 10, la suivante pour la même plantation ne sera pas le 11 ni 12.")

# --- Sélecteur calendrier (intervalle) ---
use_calendar = st.sidebar.checkbox("Choisir l'intervalle via calendrier", value=True)
start_day = 1
end_day = 31
range_start = None
range_end = None

def month_num(label):
    months = ["","Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    return months.index(label)

if use_calendar and mois_label != "(Toute l'année)":
    m = month_num(mois_label)
    days_in = calendar.monthrange(annee, m)[1]
    min_dt = date(annee, m, 1)
    max_dt = date(annee, m, days_in)
    dr = st.sidebar.date_input("Intervalle autorisé", (min_dt, max_dt), min_value=min_dt, max_value=max_dt, format="DD/MM/YYYY")
    if isinstance(dr, tuple) and len(dr) == 2:
        range_start, range_end = dr
        start_day, end_day = range_start.day, range_end.day
else:
    start_day = st.sidebar.number_input("Jour début autorisé", min_value=1, max_value=31, value=1, step=1, disabled=use_calendar and mois_label!="(Toute l'année)")
    end_day = st.sidebar.number_input("Jour fin autorisé", min_value=1, max_value=31, value=28, step=1, disabled=use_calendar and mois_label!="(Toute l'année)")

st.sidebar.header("6) Cap & tolérance")
daily_cap = st.sidebar.number_input("Plafond journalier global (kg) — 0 = désactivé", min_value=0.0, value=0.0, step=100.0)
tol_pct = st.sidebar.slider("Tolérance anti-dépassement du reste mensuel/annuel (%)", min_value=0, max_value=50, value=15, step=5)

st.sidebar.header("7) Coopérative (fallback)")
coop_input = st.sidebar.text_input("Nom de la coopérative (si colonne absente)", value="COOP Bahari")

st.sidebar.header("8) Aléa & post-traitement")
alpha = st.sidebar.slider("Dirichlet α (répartition intra-plantation)", min_value=0.5, max_value=5.0, value=1.5, step=0.5)
seed = st.sidebar.number_input("Seed aléatoire", value=42, step=1)
order_mode = st.sidebar.selectbox("Ordre du plan", ["Par date (intercalé)", "Aléatoire", "Groupé par plantation"])
auto_book = st.sidebar.checkbox("Enregistrer directement le plan comme 'livré' (historique)", value=False)

# -------------------- Exclusions (optionnel) --------------------
st.sidebar.header("9) Exclusions de plantations")
excl_file = st.sidebar.file_uploader("Excel des plantations à exclure (colonne CODE_PLANTATION)", type=["xlsx"], key="excl_file")
excl_manual = st.sidebar.text_area("Codes à exclure (coller, un par ligne)", value="")
excl_codes = set()

def detect_code_columns(cols):
    cands, seen = [], set()
    for c in cols:
        cu = str(c).strip().upper()
        if cu in {"CODE_PLANTATION", "CODE", "PLANTATION_CODE", "CODEPLANTATION"} or "CODE" in cu or "PLANT" in cu:
            if c not in seen:
                cands.append(c); seen.add(c)
    return cands

if excl_file is not None:
    try:
        df_ex = pd.read_excel(excl_file)
        if df_ex.empty:
            st.sidebar.warning("Fichier d'exclusion vide.")
        else:
            candidates = detect_code_columns(df_ex.columns)
            if not candidates:
                st.sidebar.error("Impossible de détecter la colonne code dans le fichier d'exclusion. Nommez-la CODE_PLANTATION.")
            else:
                col_sel = st.sidebar.selectbox("Colonne codes dans le fichier d'exclusion", candidates, index=0, key="excl_col")
                excl_codes |= set(df_ex[col_sel].astype(str).str.strip().unique().tolist())
    except Exception as e:
        st.sidebar.error(f"Erreur lecture exclusion: {e}")

if excl_manual.strip():
    manual_codes = [x.strip() for x in excl_manual.splitlines() if x.strip()]
    excl_codes |= set(manual_codes)

if excl_codes:
    st.sidebar.success(f"{len(excl_codes)} code(s) à exclure pris en compte.")

# -------------------- Chargement du fichier --------------------
df = None
if uploaded is not None:
    try:
        df = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Erreur de lecture du fichier: {e}")
else:
    st.info("Charge ton fichier Excel pour continuer.")

if df is not None:
    # --- Colonnes requises: CODE_PLANTATION + CODE + RENDEMENT ; NOM optionnel ; COOPÉRATIVE optionnelle
    cols_upper = {str(c).strip().upper(): c for c in df.columns}

    missing = []
    for key in ["CODE_PLANTATION", "CODE", "RENDEMENT"]:
        if key not in cols_upper:
            missing.append(key)
    if missing:
        st.error(f"Colonnes manquantes : {missing}. Attendus: CODE_PLANTATION, CODE, RENDEMENT.")
        st.stop()

    codep_col = cols_upper["CODE_PLANTATION"]
    code_col  = cols_upper["CODE"]
    # RENDEMENT
    if "RENDEMENT" in cols_upper:
        rend_col = cols_upper["RENDEMENT"]
    else:
        alt = None
        for c in df.columns:
            if str(c).strip().upper() in {"RENDEMENT (KG)","RENDEMENT KG","RENDEMENT_ANNUEL","YIELD","ANNUAL_YIELD"}:
                alt = c; break
        if alt is None:
            st.error("Colonne RENDEMENT introuvable.")
            st.stop()
        rend_col = alt

    name_col = cols_upper.get("NOM", None)
    coop_col = None
    for key in ["COOPÉRATIVE","COOPERATIVE","COOP","COOPERATIVE_NAME","COOPERATIVE_NOM"]:
        if key in cols_upper: coop_col = cols_upper[key]; break

    keep = [codep_col, code_col, rend_col] + ([name_col] if name_col else []) + ([coop_col] if coop_col else [])
    df = df[keep].copy()
    df.rename(columns={codep_col:"CODE_PLANTATION", code_col:"CODE", rend_col:"RENDEMENT"}, inplace=True)
    if name_col:
        df.rename(columns={name_col:"NOM"}, inplace=True)
    else:
        df["NOM"] = df["CODE"].astype(str)
    if not coop_col:
        df["COOPÉRATIVE"] = coop_input
    else:
        df.rename(columns={coop_col:"COOPÉRATIVE"}, inplace=True)

    df["RENDEMENT"] = pd.to_numeric(df["RENDEMENT"], errors="coerce").fillna(0)
    df = df[df["RENDEMENT"] > 0].reset_index(drop=True)

    # Appliquer exclusions par CODE_PLANTATION
    removed = 0
    if excl_codes:
        before = len(df)
        df = df[~df["CODE_PLANTATION"].astype(str).isin(excl_codes)].reset_index(drop=True)
        removed = before - len(df)
        st.warning(f"{removed} plantation(s) exclue(s) du plan.") if removed > 0 else st.info("Aucune plantation du fichier source ne correspond aux exclusions.")

    st.subheader("Aperçu des plantations (après exclusions)")
    st.dataframe(df.head(80), use_container_width=True)

    # -------------------- Cœur logique --------------------
    big_months = {1,2,3,10,11,12}
    def month_weight(m): return 0.60/6 if m in big_months else 0.40/6

    def split_by_frequency(total, k, alpha_=1.5, seed_=42):
        if k <= 1: return [float(total)]
        rng = np.random.default_rng(int(seed_))
        p = rng.dirichlet([alpha_] * k)
        return list((p * total).astype(float))

    def select_subset_greedy(cand, target):
        if target <= 0 or cand.empty: return cand.copy(), float(0)
        c = cand.sort_values("Reste mois (kg)", ascending=False).copy()
        total = 0.0; idxs = []
        for i, row in c.iterrows():
            total += float(row["Reste mois (kg)"]); idxs.append(i)
            if total >= target: break
        sub = c.loc[idxs].copy()
        return sub, total

    def select_subset_weighted(cand, target, seed_=42):
        if target <= 0 or cand.empty: return cand.copy(), float(0)
        c = cand.copy()
        weights = c["Reste mois (kg)"].to_numpy()
        if weights.sum() <= 0: return c.head(0).copy(), 0.0
        rng = np.random.default_rng(int(seed_))
        chosen, covered = [], 0.0
        remaining_idx = c.index.to_list()
        while len(remaining_idx) > 0 and covered < target:
            w = c.loc[remaining_idx, "Reste mois (kg)"].to_numpy()
            if w.sum() <= 0: break
            probs = w / w.sum()
            pick_pos = rng.choice(len(remaining_idx), replace=False, p=probs)
            pick_idx = remaining_idx.pop(int(pick_pos))
            chosen.append(pick_idx)
            covered += float(c.loc[pick_idx, "Reste mois (kg)"])
        sub = c.loc[chosen].copy()
        return sub, covered

    def allocate_target(cand, target):
        if target <= 0 or cand.empty: cand["Cible mois (kg)"] = 0.0; return cand
        cap = cand["Reste mois (kg)"].sum()
        if cap <= 0: cand["Cible mois (kg)"] = 0.0; return cand
        ratio = min(1.0, float(target)/float(cap))
        cand["Cible mois (kg)"] = cand["Reste mois (kg)"] * ratio
        return cand

    def allocate_target_greedy(cand, target):
        if target <= 0 or cand.empty: cand["Cible mois (kg)"] = 0.0; return cand
        left = float(target); out = []
        for _, row in cand.iterrows():
            r = float(row["Reste mois (kg)"])
            take = min(r, left); out.append(take); left -= take
        while len(out) < len(cand): out.append(0.0)
        cand = cand.copy(); cand["Cible mois (kg)"] = out[:len(cand)]; return cand

    def choose_valid_day(year, month, qty, candidates, per_day_totals, assigned_days_for_plant,
                         daily_cap, avoid_weekend, min_gap_days, seed_=42):
        rng = random.Random(int(seed_))
        shuffled = candidates[:]; rng.shuffle(shuffled)
        for d in shuffled:
            if avoid_weekend and date(year, month, d).weekday() >= 5: continue
            if any(abs(d - prev) < int(min_gap_days) for prev in assigned_days_for_plant): continue
            if daily_cap and daily_cap > 0 and per_day_totals.get(d, 0.0) + qty > daily_cap: continue
            return d, False
        # fallback — ignorer cap, garder règles restantes
        filt = []
        for d in candidates:
            if avoid_weekend and date(year, month, d).weekday() >= 5: continue
            if any(abs(d - prev) < int(min_gap_days) for prev in assigned_days_for_plant): continue
            filt.append(d)
        target_list = filt if filt else (candidates if candidates else [])
        if not target_list: return None, False
        best = min(target_list, key=lambda dd: per_day_totals.get(dd, 0.0))
        exceeded = (daily_cap > 0 and per_day_totals.get(best, 0.0) + qty > daily_cap)
        return best, exceeded

    def plan_for_single_month(df_in, year, month_num, freq_global, freq_rand=False, freq_min=1, freq_max=3,
                              trace_global_qty=None, mode="Proportionnel (par défaut)", seuil_min_kg=0.0,
                              order_mode="Par date (intercalé)", alpha_=1.5, use_history=True, seed_=42,
                              avoid_weekend=True, min_gap_days=0, daily_cap=0.0, start_day=1, end_day=31, tol_pct=15):
        days_in = calendar.monthrange(year, month_num)[1]
        w = month_weight(month_num)

        base = df_in.copy()
        base["Plan mensuel (kg)"] = base["RENDEMENT"] * w

        delivered_month = load_delivered_by_month(year, month_num) if use_history else {}
        delivered_year = load_delivered_by_year(year) if use_history else {}

        base["Déjà livré (mois) (kg)"] = base["CODE_PLANTATION"].astype(str).map(delivered_month).fillna(0.0)
        base["Déjà livré (annuel) (kg)"] = base["CODE_PLANTATION"].astype(str).map(delivered_year).fillna(0.0)
        base["Reste mois (kg)"]   = (base["Plan mensuel (kg)"] - base["Déjà livré (mois) (kg)"]).clip(lower=0)
        base["Reste annuel (kg)"] = (base["RENDEMENT"]        - base["Déjà livré (annuel) (kg)"]).clip(lower=0)

        candidates = base[base["Reste mois (kg)"] > 0].copy() if use_history else base.copy()
        if seuil_min_kg and seuil_min_kg > 0:
            candidates = candidates[candidates["Reste mois (kg)"] >= float(seuil_min_kg)].copy()
        if candidates.empty:
            candidates = base.copy(); candidates["Reste mois (kg)"] = candidates["Plan mensuel (kg)"]

        # Traçage global (optionnel)
        if trace_global_qty and trace_global_qty > 0:
            T = float(trace_global_qty)
            if mode == "Sous-ensemble minimal (greedy)":
                subset, covered = select_subset_greedy(candidates, T)
                cand_alloc = allocate_target_greedy(subset, min(T, covered))
            elif mode == "Échantillonnage pondéré":
                subset, covered = select_subset_weighted(candidates, T, seed_=seed_)
                cand_alloc = allocate_target(subset, min(T, covered if covered>0 else T))
            else:
                cand_alloc = allocate_target(candidates, T)
        else:
            cand_alloc = candidates.copy()
            cand_alloc["Cible mois (kg)"] = cand_alloc["Reste mois (kg)"] if use_history else cand_alloc["Plan mensuel (kg)"]

        # Tolérance (anti-dépassement)
        tol = float(tol_pct)/100.0
        cand_alloc["Cap mensuel (kg)"] = cand_alloc["Plan mensuel (kg)"] * (1.0 + tol)
        cand_alloc["Cap annuel (kg)"]  = cand_alloc["Reste annuel (kg)"]  * (1.0 + tol)
        cand_alloc["Cible mois (kg)"]  = np.minimum(cand_alloc["Cible mois (kg)"],
                                                    np.minimum(cand_alloc["Cap mensuel (kg)"], cand_alloc["Cap annuel (kg)"]))

        # Fenêtre jours
        lo = max(1, int(start_day)); hi = min(days_in, int(end_day))
        if lo > hi: lo, hi = 1, days_in
        window_days = list(range(lo, hi+1))

        # Construction des lignes
        rows = []
        rng = np.random.default_rng(int(seed_))
        rseed = int(seed_)
        per_day_totals = {}
        per_plant_days = {}

        for _, r in cand_alloc.iterrows():
            codep = str(r["CODE_PLANTATION"]); code = str(r["CODE"]); nom = str(r["NOM"]); coop = str(r["COOPÉRATIVE"])
            annual = float(r["RENDEMENT"])
            monthly_target = float(r["Cible mois (kg)"])
            if monthly_target <= 0: continue

            if freq_rand: k = int(rng.integers(int(freq_min), int(freq_max)+1))
            else: k = int(freq_global)
            k = max(1, min(k, len(window_days)))

            parts = split_by_frequency(monthly_target, k, alpha_=alpha_, seed_=rseed)
            rseed += 1

            per_plant_days.setdefault(codep, [])
            for i, qty in enumerate(parts, start=1):
                chosen_day, exceeded = choose_valid_day(
                    year, month_num, qty, candidates=window_days, per_day_totals=per_day_totals,
                    assigned_days_for_plant=per_plant_days[codep], daily_cap=daily_cap,
                    avoid_weekend=avoid_weekend, min_gap_days=min_gap_days, seed_=rseed+i
                )
                if chosen_day is None:
                    chosen_day = random.choice(window_days)
                    exceeded = (daily_cap>0 and per_day_totals.get(chosen_day,0.0)+qty>daily_cap)

                per_day_totals[chosen_day] = per_day_totals.get(chosen_day, 0.0) + float(qty)
                per_plant_days[codep].append(chosen_day)

                rows.append({
                    "Code plantation": codep,
                    "Code": code,
                    "Nom plantation": nom,
                    "Coopérative": coop,
                    "Année": year,
                    "Mois": month_num,
                    "N° évènement": i,
                    "Date planifiée": date(year, month_num, chosen_day),
                    "Rendement annuel (kg)": round(annual,3),
                    "Plan mensuel (kg)": round(float(r["Plan mensuel (kg)"]),3),
                    "Quantité prévue (kg)": round(float(qty),3),
                    "Alerte cap": "DEPASSEMENT CAP" if exceeded else "",
                })

        plan = pd.DataFrame(rows)

        if not plan.empty:
            if order_mode == "Par date (intercalé)":
                rb = np.random.default_rng(int(seed_)).random(len(plan))
                plan["_rb"] = rb
                plan = plan.sort_values(["Date planifiée","_rb"]).drop(columns=["_rb"]).reset_index(drop=True)
            elif order_mode == "Aléatoire":
                plan = plan.sample(frac=1.0, random_state=int(seed_)).reset_index(drop=True)

        return plan

    # -------------------- Génération --------------------
    st.header("Génération du plan")
    generate_clicked = st.button("🚀 Générer le plan")
    if generate_clicked:
        if mois_label == "(Toute l'année)":
            frames = []
            for m in range(1,13):
                days_in = calendar.monthrange(annee, m)[1]
                lo = 1; hi = days_in
                if use_calendar and isinstance(range_start, date) and isinstance(range_end, date):
                    if date(annee, m, days_in) < range_start or date(annee, m, 1) > range_end:
                        continue
                    lo = 1 if m != range_start.month else max(1, range_start.day)
                    hi = days_in if m != range_end.month else min(days_in, range_end.day)
                else:
                    lo = start_day; hi = end_day if end_day <= days_in else days_in

                frames.append(
                    plan_for_single_month(
                        df, annee, m, freq_global,
                        freq_rand=freq_rand, freq_min=freq_min, freq_max=freq_max,
                        trace_global_qty=quota_global if trace_global else None,
                        mode=mode, seuil_min_kg=seuil_min,
                        order_mode=order_mode, alpha_=alpha, use_history=True,
                        seed_=seed+m, avoid_weekend=avoid_weekend, min_gap_days=min_gap_days,
                        daily_cap=daily_cap, start_day=lo, end_day=hi, tol_pct=tol_pct
                    )
                )
            plan = pd.concat(frames, ignore_index=True) if len(frames) else pd.DataFrame([])
        else:
            m = month_num(mois_label)
            days_in = calendar.monthrange(annee, m)[1]
            lo = start_day; hi = end_day if end_day <= days_in else days_in
            if use_calendar and isinstance(range_start, date) and isinstance(range_end, date):
                lo = range_start.day; hi = min(days_in, range_end.day)

            plan = plan_for_single_month(
                df, annee, m, freq_global,
                freq_rand=freq_rand, freq_min=freq_min, freq_max=freq_max,
                trace_global_qty=quota_global if trace_global else None,
                mode=mode, seuil_min_kg=seuil_min,
                order_mode=order_mode, alpha_=alpha, use_history=True,
                seed_=seed, avoid_weekend=avoid_weekend, min_gap_days=min_gap_days,
                daily_cap=daily_cap, start_day=lo, end_day=hi, tol_pct=tol_pct
            )

        if plan.empty:
            st.warning("Aucun plan généré (vérifie capacité, historique et paramètres).")
        else:
            monthly_summary = (
                plan.groupby(["Code plantation","Code","Nom plantation","Coopérative","Année","Mois"], as_index=False)
                    .agg(**{
                        "Plan mensuel (kg)": ("Plan mensuel (kg)", "max"),
                        "Prévu total mois (kg)": ("Quantité prévue (kg)","sum")
                    })
            )

            st.success("Plan généré ✅")
            st.subheader("Aperçu du plan (ordre choisi)")
            st.dataframe(plan.head(200), use_container_width=True, height=420)

            st.subheader("Résumé mensuel")
            st.dataframe(monthly_summary.head(100), use_container_width=True)

            # Export Excel (Plan + Suivi + Résumé + Paramètres)
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="xlsxwriter", datetime_format="yyyy-mm-dd", date_format="yyyy-mm-dd") as writer:
                plan.to_excel(writer, sheet_name="Plan", index=False)

                tracking = plan.copy()
                tracking["Quantité livrée (kg)"] = ""
                tracking["Livré à date (kg)"] = 0.0
                tracking["Livré annuel (kg)"] = 0.0
                tracking["Reste annuel (kg)"] = 0.0
                tracking["Livré du mois (kg)"] = 0.0
                tracking["Reste du mois (kg)"] = 0.0
                tracking["Alerte annuel"] = ""
                cols = ["Code plantation","Code","Nom plantation","Coopérative","Année","Mois","N° évènement","Date planifiée",
                        "Rendement annuel (kg)","Plan mensuel (kg)","Quantité prévue (kg)","Alerte cap",
                        "Quantité livrée (kg)","Livré à date (kg)","Livré annuel (kg)","Reste annuel (kg)",
                        "Livré du mois (kg)","Reste du mois (kg)","Alerte annuel"]
                tracking = tracking[cols]
                tracking.to_excel(writer, sheet_name="Suivi", index=False)
                ws = writer.sheets["Suivi"]
                n = len(tracking)
                for i in range(2, n+2):
                    ws.write_formula(i-1, 13, f'=IF($A{i}="","", SUMIFS($M:$M,$A:$A,$A{i}))')  # Livré annuel
                    ws.write_formula(i-1, 14, f'=IF($A{i}="","", $I{i} - $N{i})')              # Reste annuel
                    ws.write_formula(i-1, 15, f'=IF($A{i}="","", SUMPRODUCT(($A:$A=$A{i})*(MONTH($H:$H)=MONTH($H{i}))*(YEAR($H:$H)=YEAR($H{i}))*$M:$M))')  # Livré du mois
                    ws.write_formula(i-1, 16, f'=IF($A{i}="","", $J{i} - $O{i})')              # Reste du mois
                    ws.write_formula(i-1, 17, f'=IF($L{i}<0,"EXCÈS","OK")')                    # Alerte annuel

                monthly_summary.to_excel(writer, sheet_name="Résumé mensuel", index=False)
                params = pd.DataFrame({
                    "Paramètre": [
                        "Année","Mois","Fréquence globale","Fréq. aléatoire","Plage fréq.",
                        "Target actif","Quantité à tracer (kg)","Mode répartition","Seuil min (kg)",
                        "Éviter week-ends","Min gap (jours)","Fenêtre via calendrier","Cap journalier (kg)",
                        "Tolérance dépassement (%)","Dirichlet α","Historique activé","Seed","Ordre",
                        "Nb exclusions"
                    ],
                    "Valeur": [
                        annee, mois_label, freq_global, "Oui" if freq_rand else "Non", f"{freq_min}-{freq_max}",
                        "Oui" if trace_global else "Non", quota_global if trace_global else "", mode, seuil_min,
                        "Oui" if avoid_weekend else "Non", min_gap_days, "Oui" if use_calendar else "Non", daily_cap,
                        tol_pct, alpha, "Oui", seed, order_mode, len(excl_codes)
                    ]
                })
                params.to_excel(writer, sheet_name="Paramètres", index=False)

            st.download_button("💾 Télécharger l'Excel (Plan + Suivi + Résumé + Paramètres)",
                               data=buffer.getvalue(),
                               file_name=f"plan_livraisons_{annee}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            if auto_book:
                count = 0
                for _, rr in plan.iterrows():
                    if insert_delivery(rr["Code plantation"], rr["Nom plantation"], rr["Coopérative"], int(rr["Année"]), int(rr["Mois"]), rr["Date planifiée"], float(rr["Quantité prévue (kg)"])):
                        count += 1
                st.success(f"{count} lignes enregistrées comme 'livrées' dans l'historique.")

    # -------------------- Historique --------------------
    st.header("Historique des livraisons (base de données)")
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Ajouter une livraison (saisie rapide)")
        if df is not None and 'CODE_PLANTATION' in df.columns:
            # menu affiche CODE + NOM + COOP
            df_sel = df.copy()
            df_sel["__label__"] = df_sel["CODE_PLANTATION"].astype(str) + " | " + df_sel["CODE"].astype(str) + " | " + df_sel["NOM"].astype(str)
            label = st.selectbox("Choisir la plantation", df_sel["__label__"].tolist())
            row = df_sel[df_sel["__label__"]==label].iloc[0]
            code_sel = row["CODE_PLANTATION"]; nom_sel = row["NOM"]; coop_sel = row["COOPÉRATIVE"]
        else:
            code_sel = st.text_input("Code plantation (manuel)")
            nom_sel  = st.text_input("Nom plantation (manuel)")
            coop_sel = st.text_input("Coopérative (manuel)", value=coop_input)
        annee_in = st.number_input("Année livraison", min_value=2020, max_value=2035, value=int(annee))
        mois_in = st.selectbox("Mois livraison", ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"], index=0)
        mois_num_in = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"].index(mois_in)+1
        jour = st.number_input("Jour livraison", min_value=1, max_value=31, value=1)
        qty_in = st.number_input("Quantité livrée (kg)", min_value=0.0, value=0.0, step=10.0)
        if st.button("➕ Enregistrer la livraison"):
            try:
                dte = date(int(annee_in), int(mois_num_in), int(jour))
                if insert_delivery(code_sel, nom_sel, coop_sel, annee_in, mois_num_in, dte, qty_in):
                    st.success("Livraison enregistrée.")
                else:
                    st.info("Cette livraison existe déjà (même code/date/quantité).")
            except Exception as e:
                st.error(f"Erreur enregistrement : {e}")

    with col2:
        st.subheader("Importer un Excel 'Suivi'")
        upl_hist = st.file_uploader("Sélectionner un 'Suivi' exporté depuis l'outil", type=["xlsx"], key="import_hist")
        if upl_hist is not None:
            try:
                tdf = pd.read_excel(upl_hist, sheet_name="Suivi")
                needed = ["Code plantation","Nom plantation","Coopérative","Année","Mois","Date planifiée","Quantité livrée (kg)"]
                missing2 = [c for c in needed if c not in tdf.columns]
                if missing2:
                    st.error(f"Colonnes manquantes dans 'Suivi' : {missing2}")
                else:
                    tdf = tdf[pd.to_numeric(tdf["Quantité livrée (kg)"], errors="coerce").fillna(0) > 0]
                    cnt = 0
                    for _, rr in tdf.iterrows():
                        if insert_delivery(rr["Code plantation"], rr["Nom plantation"], rr["Coopérative"], int(rr["Année"]), int(rr["Mois"]), rr["Date planifiée"], float(rr["Quantité livrée (kg)"])):
                            cnt += 1
                    st.success(f"{cnt} lignes importées dans l'historique.")
            except Exception as e:
                st.error(f"Erreur import : {e}")

    st.subheader("Gestion avancée de l'historique")
    hcol1, hcol2, hcol3 = st.columns(3)
    with hcol1:
        if st.button("📤 Exporter l'historique (CSV)"):
            hist = fetch_history(limit=1000000)
            csv = hist.to_csv(index=False).encode("utf-8")
            st.download_button("Télécharger CSV", data=csv, file_name="historique_livraisons.csv", mime="text/csv", key="dl_csv")
    with hcol2:
        if st.button("📤 Exporter l'historique (Excel)"):
            hist = fetch_history(limit=1000000)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
                hist.to_excel(w, sheet_name="Historique", index=False)
            st.download_button("Télécharger Excel", data=buf.getvalue(), file_name="historique_livraisons.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xlsx")
    with hcol3:
        confirm = st.text_input("Écrire SUPPRIMER pour vider l'historique")
        if st.button("🗑️ Vider l'historique"):
            if confirm.strip().upper() == "SUPPRIMER":
                clear_history()
                st.success("Historique vidé.")
            else:
                st.warning("Confirmation manquante : tape 'SUPPRIMER'.")

# ==================== 10) FICHE D'EXPORT ====================
st.header("Fiche d'export — remplir le modèle à partir du plan")
with st.expander("Paramètres de la fiche d'export", expanded=True):
    colA, colB, colC = st.columns(3)
    with colA:
        exportateur_nom = st.text_input("Nom de l'exportateur", value="Bahari S.A.R.L.")
        export_lot_no = st.text_input("Export Lot N°", value="LOT-2025-001")
    with colB:
        connaissement = st.text_input("Connaissement (B/L)", value="BL-123456")
        destination = st.text_input("Destination (port/pays)", value="Abidjan, CI")
    with colC:
        contact = st.text_input("Contact / Référence", value="Opérations Export")
        commentaire = st.text_area("Commentaire (facultatif)", value="", height=95)
    tpl_file = st.file_uploader("Modèle Excel à compléter (facultatif)", type=["xlsx"], key="export_tpl")

def build_export_workbook(plan_df, meta, template_file=None):
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    import io
    if template_file is not None:
        try:
            wb = load_workbook(template_file)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()
    if "EXPORT_META" in wb.sheetnames: wb.remove(wb["EXPORT_META"])
    ws_meta = wb.create_sheet("EXPORT_META")
    ws_meta.append(["Clé","Valeur"])
    for k,v in [
        ("Exportateur", meta.get("exportateur_nom","")),
        ("Export Lot N°", meta.get("export_lot_no","")),
        ("Connaissement", meta.get("connaissement","")),
        ("Destination", meta.get("destination","")),
        ("Contact", meta.get("contact","")),
        ("Commentaire", meta.get("commentaire","")),
        ("Généré le", pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
    ]:
        ws_meta.append([k,v])

    if "PLANNING_EXPORT" in wb.sheetnames: wb.remove(wb["PLANNING_EXPORT"])
    ws_det = wb.create_sheet("PLANNING_EXPORT")
    cols = ["Export Lot N°","Connaissement","Exportateur","Année","Mois","N° évènement","Date planifiée",
            "Code plantation","Code","Nom plantation","Coopérative","Quantité prévue (kg)","Plan mensuel (kg)","Rendement annuel (kg)"]
    ws_det.append(cols)
    if len(plan_df) > 0:
        for _, r in plan_df.iterrows():
            ws_det.append([
                meta.get("export_lot_no",""), meta.get("connaissement",""), meta.get("exportateur_nom",""),
                int(r.get("Année","") or 0), int(r.get("Mois","") or 0), int(r.get("N° évènement",1) or 1),
                str(pd.to_datetime(r.get("Date planifiée")).date()) if r.get("Date planifiée","")!="" else "",
                str(r.get("Code plantation","")), str(r.get("Code","")), str(r.get("Nom plantation","")),
                str(r.get("Coopérative","")), float(r.get("Quantité prévue (kg)",0.0)), float(r.get("Plan mensuel (kg)",0.0)),
                float(r.get("Rendement annuel (kg)",0.0)),
            ])
    if "RESUME_JOUR" in wb.sheetnames: wb.remove(wb["RESUME_JOUR"])
    ws_sum = wb.create_sheet("RESUME_JOUR")
    if len(plan_df) > 0:
        pivot = (plan_df.groupby("Date planifiée", as_index=False)["Quantité prévue (kg)"].sum().sort_values("Date planifiée"))
        ws_sum.append(["Date","Total prévu (kg)"])
        for _, rr in pivot.iterrows():
            ws_sum.append([str(pd.to_datetime(rr["Date planifiée"]).date()), float(rr["Quantité prévue (kg)"])])

    for ws in [ws_meta, ws_det, ws_sum]:
        for col in range(1, ws.max_column+1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col)].width = 18

    out = io.BytesIO(); wb.save(out); return out.getvalue()

if 'plan' in locals() and isinstance(plan, pd.DataFrame) and not plan.empty:
    meta = {"exportateur_nom": exportateur_nom, "export_lot_no": export_lot_no, "connaissement": connaissement,
            "destination": destination, "contact": contact, "commentaire": commentaire}
    if st.button("📄 Construire la fiche d'export à partir du plan"):
        file_bytes = build_export_workbook(plan, meta, tpl_file)
        st.download_button("💾 Télécharger la fiche d'export (.xlsx)", data=file_bytes,
                           file_name=f"FICHE_EXPORT_{export_lot_no or 'LOT'}_{annee}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("👉 Génère d'abord un plan pour activer la fiche d'export.")

import io, os, sqlite3, calendar, random
from datetime import date, datetime
from pathlib import Path
import numpy as np
import pandas as pd
import streamlit as st

# -------------------- Page --------------------
st.set_page_config(page_title="Planificateur des livraisons (rendement annuel)", layout="wide")
st.title("📦 Planificateur des livraisons par plantation — version avancée")
st.caption("Saisonnalité 60/40, Target, fréquences aléatoires, ordre intercalé, calendrier & cap, tolérance, exclusions, historique SQLite, fiche d’export.")

DB_PATH = str(Path(__file__).resolve().parent / "state.sqlite")

# -------------------- Base SQLite --------------------
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
        CREATE TABLE IF NOT EXISTS livraisons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code_plantation TEXT NOT NULL,
            nom_plantation TEXT,
            cooperative TEXT,
            annee INTEGER NOT NULL,
            mois INTEGER NOT NULL,
            date_livraison TEXT NOT NULL,
            quantite_kg REAL NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        c.execute("CREATE INDEX IF NOT EXISTS idx_livr_ym_code ON livraisons(annee, mois, code_plantation)")
        conn.commit()

def exists_delivery(code, annee, mois, dte, qty):
    with sqlite3.connect(DB_PATH) as conn:
        df = pd.read_sql_query(
            "SELECT COUNT(*) n FROM livraisons WHERE code_plantation=? AND annee=? AND mois=? AND date_livraison=? AND quantite_kg=?",
            conn, params=(str(code), int(annee), int(mois), str(dte), float(qty)))
    return int(df.iloc[0]["n"]) > 0

def insert_delivery(code, nom, coop, annee, mois, dte, qty):
    if exists_delivery(code, annee, mois, dte, qty):
        return False
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""INSERT INTO livraisons (code_plantation, nom_plantation, cooperative, annee, mois, date_livraison, quantite_kg)
                        VALUES (?,?,?,?,?,?,?)""", (str(code), nom, coop, int(annee), int(mois), str(dte), float(qty)))
        conn.commit()
    return True

def load_delivered_by_month(annee, mois):
    with sqlite3.connect(DB_PATH) as conn:
        q = "SELECT code_plantation, SUM(quantite_kg) sumq FROM livraisons WHERE annee=? AND mois=? GROUP BY code_plantation"
        df = pd.read_sql_query(q, conn, params=(int(annee), int(mois)))
    return dict(zip(df["code_plantation"], df["sumq"])) if len(df) else {}

def load_delivered_by_year(annee):
    with sqlite3.connect(DB_PATH) as conn:
        q = "SELECT code_plantation, SUM(quantite_kg) sumq FROM livraisons WHERE annee=? GROUP BY code_plantation"
        df = pd.read_sql_query(q, conn, params=(int(annee),))
    return dict(zip(df["code_plantation"], df["sumq"])) if len(df) else {}

def fetch_history(limit=200000):
    with sqlite3.connect(DB_PATH) as conn:
        df = pd.read_sql_query("SELECT * FROM livraisons ORDER BY created_at DESC LIMIT ?", conn, params=(int(limit),))
    return df

def clear_history():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("DELETE FROM livraisons")
        conn.commit()

init_db()

# -------------------- UI (barre latérale) --------------------
st.sidebar.header("1) Fichier source plantations")
st.sidebar.write("**Excel requis :** `CODE_PLANTATION`, `CODE`, `RENDEMENT`  — `NOM` **optionnel** (si absent ⇒ `NOM = CODE`)  — `COOPÉRATIVE` optionnel.")
uploaded = st.sidebar.file_uploader("Glisser un .xlsx (200 MB max)", type=["xlsx"])

st.sidebar.header("2) Période & fréquence")
annee = st.sidebar.number_input("Année", min_value=2020, max_value=2035, value=2025, step=1)
mois_label = st.sidebar.selectbox("Mois précis",
                                  ["(Toute l'année)","Janvier","Février","Mars","Avril","Mai","Juin",
                                   "Juillet","Août","Septembre","Octobre","Novembre","Décembre"], index=0)
freq_global = st.sidebar.slider("Fréquence globale (fallback)", min_value=1, max_value=5, value=2, step=1)

st.sidebar.header("3) Fréquence aléatoire par plantation")
freq_rand = st.sidebar.checkbox("Activer fréquence aléatoire (par plantation)", value=True)
minmax = st.sidebar.slider("Plage de fréquence (livraisons / mois)", 1, 5, (1, 3), disabled=not freq_rand)
freq_min, freq_max = minmax if freq_rand else (freq_global, freq_global)

st.sidebar.header("4) Target / Traçage global (optionnel)")
trace_global = st.sidebar.checkbox("Activer 'Quantité à tracer' (kg)")
quota_global = st.sidebar.number_input("Quantité à tracer (kg)", min_value=0.0, value=24000.0, step=100.0, disabled=not trace_global)
mode = st.sidebar.selectbox("Mode répartition Target", ["Proportionnel (par défaut)", "Sous-ensemble minimal (greedy)", "Échantillonnage pondéré"])
seuil_min = st.sidebar.number_input("Seuil min (kg) par plantation", min_value=0.0, value=0.0, step=50.0)

st.sidebar.header("5) Contraintes calendrier")
avoid_weekend = st.sidebar.checkbox("Éviter les week-ends (samedi/dimanche)", value=True)
min_gap_days = st.sidebar.number_input("Écart minimum entre deux livraisons d'une même plantation (jours)", min_value=0, max_value=15, value=3, step=1)
st.sidebar.caption("Ex. 3 jours: si une livraison est le 10, la suivante pour la même plantation ne sera pas le 11 ni 12.")

# --- Sélecteur calendrier (intervalle) ---
use_calendar = st.sidebar.checkbox("Choisir l'intervalle via calendrier", value=True)
start_day = 1
end_day = 31
range_start = None
range_end = None

def month_num(label):
    months = ["","Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    return months.index(label)

if use_calendar and mois_label != "(Toute l'année)":
    m = month_num(mois_label)
    days_in = calendar.monthrange(annee, m)[1]
    min_dt = date(annee, m, 1)
    max_dt = date(annee, m, days_in)
    dr = st.sidebar.date_input("Intervalle autorisé", (min_dt, max_dt), min_value=min_dt, max_value=max_dt, format="DD/MM/YYYY")
    if isinstance(dr, tuple) and len(dr) == 2:
        range_start, range_end = dr
        start_day, end_day = range_start.day, range_end.day
else:
    start_day = st.sidebar.number_input("Jour début autorisé", min_value=1, max_value=31, value=1, step=1, disabled=use_calendar and mois_label!="(Toute l'année)")
    end_day = st.sidebar.number_input("Jour fin autorisé", min_value=1, max_value=31, value=28, step=1, disabled=use_calendar and mois_label!="(Toute l'année)")

st.sidebar.header("6) Cap & tolérance")
daily_cap = st.sidebar.number_input("Plafond journalier global (kg) — 0 = désactivé", min_value=0.0, value=0.0, step=100.0)
tol_pct = st.sidebar.slider("Tolérance anti-dépassement du reste mensuel/annuel (%)", min_value=0, max_value=50, value=15, step=5)

st.sidebar.header("7) Coopérative (fallback)")
coop_input = st.sidebar.text_input("Nom de la coopérative (si colonne absente)", value="COOP Bahari")

st.sidebar.header("8) Aléa & post-traitement")
alpha = st.sidebar.slider("Dirichlet α (répartition intra-plantation)", min_value=0.5, max_value=5.0, value=1.5, step=0.5)
seed = st.sidebar.number_input("Seed aléatoire", value=42, step=1)
order_mode = st.sidebar.selectbox("Ordre du plan", ["Par date (intercalé)", "Aléatoire", "Groupé par plantation"])
auto_book = st.sidebar.checkbox("Enregistrer directement le plan comme 'livré' (historique)", value=False)

# -------------------- Exclusions (optionnel) --------------------
st.sidebar.header("9) Exclusions de plantations")
excl_file = st.sidebar.file_uploader("Excel des plantations à exclure (colonne CODE_PLANTATION)", type=["xlsx"], key="excl_file")
excl_manual = st.sidebar.text_area("Codes à exclure (coller, un par ligne)", value="")
excl_codes = set()

def detect_code_columns(cols):
    cands, seen = [], set()
    for c in cols:
        cu = str(c).strip().upper()
        if cu in {"CODE_PLANTATION", "CODE", "PLANTATION_CODE", "CODEPLANTATION"} or "CODE" in cu or "PLANT" in cu:
            if c not in seen:
                cands.append(c); seen.add(c)
    return cands

if excl_file is not None:
    try:
        df_ex = pd.read_excel(excl_file)
        if df_ex.empty:
            st.sidebar.warning("Fichier d'exclusion vide.")
        else:
            candidates = detect_code_columns(df_ex.columns)
            if not candidates:
                st.sidebar.error("Impossible de détecter la colonne code dans le fichier d'exclusion. Nommez-la CODE_PLANTATION.")
            else:
                col_sel = st.sidebar.selectbox("Colonne codes dans le fichier d'exclusion", candidates, index=0, key="excl_col")
                excl_codes |= set(df_ex[col_sel].astype(str).str.strip().unique().tolist())
    except Exception as e:
        st.sidebar.error(f"Erreur lecture exclusion: {e}")

if excl_manual.strip():
    manual_codes = [x.strip() for x in excl_manual.splitlines() if x.strip()]
    excl_codes |= set(manual_codes)

if excl_codes:
    st.sidebar.success(f"{len(excl_codes)} code(s) à exclure pris en compte.")

# -------------------- Chargement du fichier --------------------
df = None
if uploaded is not None:
    try:
        df = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Erreur de lecture du fichier: {e}")
else:
    st.info("Charge ton fichier Excel pour continuer.")

if df is not None:
    # --- Colonnes requises: CODE_PLANTATION + CODE + RENDEMENT ; NOM optionnel ; COOPÉRATIVE optionnelle
    cols_upper = {str(c).strip().upper(): c for c in df.columns}

    missing = []
    for key in ["CODE_PLANTATION", "CODE", "RENDEMENT"]:
        if key not in cols_upper:
            missing.append(key)
    if missing:
        st.error(f"Colonnes manquantes : {missing}. Attendus: CODE_PLANTATION, CODE, RENDEMENT.")
        st.stop()

    codep_col = cols_upper["CODE_PLANTATION"]
    code_col  = cols_upper["CODE"]
    # RENDEMENT
    if "RENDEMENT" in cols_upper:
        rend_col = cols_upper["RENDEMENT"]
    else:
        alt = None
        for c in df.columns:
            if str(c).strip().upper() in {"RENDEMENT (KG)","RENDEMENT KG","RENDEMENT_ANNUEL","YIELD","ANNUAL_YIELD"}:
                alt = c; break
        if alt is None:
            st.error("Colonne RENDEMENT introuvable.")
            st.stop()
        rend_col = alt

    name_col = cols_upper.get("NOM", None)
    coop_col = None
    for key in ["COOPÉRATIVE","COOPERATIVE","COOP","COOPERATIVE_NAME","COOPERATIVE_NOM"]:
        if key in cols_upper: coop_col = cols_upper[key]; break

    keep = [codep_col, code_col, rend_col] + ([name_col] if name_col else []) + ([coop_col] if coop_col else [])
    df = df[keep].copy()
    df.rename(columns={codep_col:"CODE_PLANTATION", code_col:"CODE", rend_col:"RENDEMENT"}, inplace=True)
    if name_col:
        df.rename(columns={name_col:"NOM"}, inplace=True)
    else:
        df["NOM"] = df["CODE"].astype(str)
    if not coop_col:
        df["COOPÉRATIVE"] = coop_input
    else:
        df.rename(columns={coop_col:"COOPÉRATIVE"}, inplace=True)

    df["RENDEMENT"] = pd.to_numeric(df["RENDEMENT"], errors="coerce").fillna(0)
    df = df[df["RENDEMENT"] > 0].reset_index(drop=True)

    # Appliquer exclusions par CODE_PLANTATION
    # Appliquer exclusions par CODE_PLANTATION
removed = 0
if excl_codes:
    before = len(df)
    # on calcule d'abord le masque, puis on filtre
    mask = df["CODE_PLANTATION"].astype(str).isin(excl_codes)
    removed = int(mask.sum())
    df = df[~mask].reset_index(drop=True)

# message d’info (bloc if/else classique, pas d’expression ternaire)
if removed > 0:
    st.warning(f"{removed} plantation(s) exclue(s) du plan.")
else:
    st.info("Aucune plantation du fichier source ne correspond aux exclusions.")


    # -------------------- Cœur logique --------------------
    big_months = {1,2,3,10,11,12}
    def month_weight(m): return 0.60/6 if m in big_months else 0.40/6

    def split_by_frequency(total, k, alpha_=1.5, seed_=42):
        if k <= 1: return [float(total)]
        rng = np.random.default_rng(int(seed_))
        p = rng.dirichlet([alpha_] * k)
        return list((p * total).astype(float))

    def select_subset_greedy(cand, target):
        if target <= 0 or cand.empty: return cand.copy(), float(0)
        c = cand.sort_values("Reste mois (kg)", ascending=False).copy()
        total = 0.0; idxs = []
        for i, row in c.iterrows():
            total += float(row["Reste mois (kg)"]); idxs.append(i)
            if total >= target: break
        sub = c.loc[idxs].copy()
        return sub, total

    def select_subset_weighted(cand, target, seed_=42):
        if target <= 0 or cand.empty: return cand.copy(), float(0)
        c = cand.copy()
        weights = c["Reste mois (kg)"].to_numpy()
        if weights.sum() <= 0: return c.head(0).copy(), 0.0
        rng = np.random.default_rng(int(seed_))
        chosen, covered = [], 0.0
        remaining_idx = c.index.to_list()
        while len(remaining_idx) > 0 and covered < target:
            w = c.loc[remaining_idx, "Reste mois (kg)"].to_numpy()
            if w.sum() <= 0: break
            probs = w / w.sum()
            pick_pos = rng.choice(len(remaining_idx), replace=False, p=probs)
            pick_idx = remaining_idx.pop(int(pick_pos))
            chosen.append(pick_idx)
            covered += float(c.loc[pick_idx, "Reste mois (kg)"])
        sub = c.loc[chosen].copy()
        return sub, covered

    def allocate_target(cand, target):
        if target <= 0 or cand.empty: cand["Cible mois (kg)"] = 0.0; return cand
        cap = cand["Reste mois (kg)"].sum()
        if cap <= 0: cand["Cible mois (kg)"] = 0.0; return cand
        ratio = min(1.0, float(target)/float(cap))
        cand["Cible mois (kg)"] = cand["Reste mois (kg)"] * ratio
        return cand

    def allocate_target_greedy(cand, target):
        if target <= 0 or cand.empty: cand["Cible mois (kg)"] = 0.0; return cand
        left = float(target); out = []
        for _, row in cand.iterrows():
            r = float(row["Reste mois (kg)"])
            take = min(r, left); out.append(take); left -= take
        while len(out) < len(cand): out.append(0.0)
        cand = cand.copy(); cand["Cible mois (kg)"] = out[:len(cand)]; return cand

    def choose_valid_day(year, month, qty, candidates, per_day_totals, assigned_days_for_plant,
                         daily_cap, avoid_weekend, min_gap_days, seed_=42):
        rng = random.Random(int(seed_))
        shuffled = candidates[:]; rng.shuffle(shuffled)
        for d in shuffled:
            if avoid_weekend and date(year, month, d).weekday() >= 5: continue
            if any(abs(d - prev) < int(min_gap_days) for prev in assigned_days_for_plant): continue
            if daily_cap and daily_cap > 0 and per_day_totals.get(d, 0.0) + qty > daily_cap: continue
            return d, False
        # fallback — ignorer cap, garder règles restantes
        filt = []
        for d in candidates:
            if avoid_weekend and date(year, month, d).weekday() >= 5: continue
            if any(abs(d - prev) < int(min_gap_days) for prev in assigned_days_for_plant): continue
            filt.append(d)
        target_list = filt if filt else (candidates if candidates else [])
        if not target_list: return None, False
        best = min(target_list, key=lambda dd: per_day_totals.get(dd, 0.0))
        exceeded = (daily_cap > 0 and per_day_totals.get(best, 0.0) + qty > daily_cap)
        return best, exceeded

    def plan_for_single_month(df_in, year, month_num, freq_global, freq_rand=False, freq_min=1, freq_max=3,
                              trace_global_qty=None, mode="Proportionnel (par défaut)", seuil_min_kg=0.0,
                              order_mode="Par date (intercalé)", alpha_=1.5, use_history=True, seed_=42,
                              avoid_weekend=True, min_gap_days=0, daily_cap=0.0, start_day=1, end_day=31, tol_pct=15):
        days_in = calendar.monthrange(year, month_num)[1]
        w = month_weight(month_num)

        base = df_in.copy()
        base["Plan mensuel (kg)"] = base["RENDEMENT"] * w

        delivered_month = load_delivered_by_month(year, month_num) if use_history else {}
        delivered_year = load_delivered_by_year(year) if use_history else {}

        base["Déjà livré (mois) (kg)"] = base["CODE_PLANTATION"].astype(str).map(delivered_month).fillna(0.0)
        base["Déjà livré (annuel) (kg)"] = base["CODE_PLANTATION"].astype(str).map(delivered_year).fillna(0.0)
        base["Reste mois (kg)"]   = (base["Plan mensuel (kg)"] - base["Déjà livré (mois) (kg)"]).clip(lower=0)
        base["Reste annuel (kg)"] = (base["RENDEMENT"]        - base["Déjà livré (annuel) (kg)"]).clip(lower=0)

        candidates = base[base["Reste mois (kg)"] > 0].copy() if use_history else base.copy()
        if seuil_min_kg and seuil_min_kg > 0:
            candidates = candidates[candidates["Reste mois (kg)"] >= float(seuil_min_kg)].copy()
        if candidates.empty:
            candidates = base.copy(); candidates["Reste mois (kg)"] = candidates["Plan mensuel (kg)"]

        # Traçage global (optionnel)
        if trace_global_qty and trace_global_qty > 0:
            T = float(trace_global_qty)
            if mode == "Sous-ensemble minimal (greedy)":
                subset, covered = select_subset_greedy(candidates, T)
                cand_alloc = allocate_target_greedy(subset, min(T, covered))
            elif mode == "Échantillonnage pondéré":
                subset, covered = select_subset_weighted(candidates, T, seed_=seed_)
                cand_alloc = allocate_target(subset, min(T, covered if covered>0 else T))
            else:
                cand_alloc = allocate_target(candidates, T)
        else:
            cand_alloc = candidates.copy()
            cand_alloc["Cible mois (kg)"] = cand_alloc["Reste mois (kg)"] if use_history else cand_alloc["Plan mensuel (kg)"]

        # Tolérance (anti-dépassement)
        tol = float(tol_pct)/100.0
        cand_alloc["Cap mensuel (kg)"] = cand_alloc["Plan mensuel (kg)"] * (1.0 + tol)
        cand_alloc["Cap annuel (kg)"]  = cand_alloc["Reste annuel (kg)"]  * (1.0 + tol)
        cand_alloc["Cible mois (kg)"]  = np.minimum(cand_alloc["Cible mois (kg)"],
                                                    np.minimum(cand_alloc["Cap mensuel (kg)"], cand_alloc["Cap annuel (kg)"]))

        # Fenêtre jours
        lo = max(1, int(start_day)); hi = min(days_in, int(end_day))
        if lo > hi: lo, hi = 1, days_in
        window_days = list(range(lo, hi+1))

        # Construction des lignes
        rows = []
        rng = np.random.default_rng(int(seed_))
        rseed = int(seed_)
        per_day_totals = {}
        per_plant_days = {}

        for _, r in cand_alloc.iterrows():
            codep = str(r["CODE_PLANTATION"]); code = str(r["CODE"]); nom = str(r["NOM"]); coop = str(r["COOPÉRATIVE"])
            annual = float(r["RENDEMENT"])
            monthly_target = float(r["Cible mois (kg)"])
            if monthly_target <= 0: continue

            if freq_rand: k = int(rng.integers(int(freq_min), int(freq_max)+1))
            else: k = int(freq_global)
            k = max(1, min(k, len(window_days)))

            parts = split_by_frequency(monthly_target, k, alpha_=alpha_, seed_=rseed)
            rseed += 1

            per_plant_days.setdefault(codep, [])
            for i, qty in enumerate(parts, start=1):
                chosen_day, exceeded = choose_valid_day(
                    year, month_num, qty, candidates=window_days, per_day_totals=per_day_totals,
                    assigned_days_for_plant=per_plant_days[codep], daily_cap=daily_cap,
                    avoid_weekend=avoid_weekend, min_gap_days=min_gap_days, seed_=rseed+i
                )
                if chosen_day is None:
                    chosen_day = random.choice(window_days)
                    exceeded = (daily_cap>0 and per_day_totals.get(chosen_day,0.0)+qty>daily_cap)

                per_day_totals[chosen_day] = per_day_totals.get(chosen_day, 0.0) + float(qty)
                per_plant_days[codep].append(chosen_day)

                rows.append({
                    "Code plantation": codep,
                    "Code": code,
                    "Nom plantation": nom,
                    "Coopérative": coop,
                    "Année": year,
                    "Mois": month_num,
                    "N° évènement": i,
                    "Date planifiée": date(year, month_num, chosen_day),
                    "Rendement annuel (kg)": round(annual,3),
                    "Plan mensuel (kg)": round(float(r["Plan mensuel (kg)"]),3),
                    "Quantité prévue (kg)": round(float(qty),3),
                    "Alerte cap": "DEPASSEMENT CAP" if exceeded else "",
                })

        plan = pd.DataFrame(rows)

        if not plan.empty:
            if order_mode == "Par date (intercalé)":
                rb = np.random.default_rng(int(seed_)).random(len(plan))
                plan["_rb"] = rb
                plan = plan.sort_values(["Date planifiée","_rb"]).drop(columns=["_rb"]).reset_index(drop=True)
            elif order_mode == "Aléatoire":
                plan = plan.sample(frac=1.0, random_state=int(seed_)).reset_index(drop=True)

        return plan

    # -------------------- Génération --------------------
    st.header("Génération du plan")
    generate_clicked = st.button("🚀 Générer le plan")
    if generate_clicked:
        if mois_label == "(Toute l'année)":
            frames = []
            for m in range(1,13):
                days_in = calendar.monthrange(annee, m)[1]
                lo = 1; hi = days_in
                if use_calendar and isinstance(range_start, date) and isinstance(range_end, date):
                    if date(annee, m, days_in) < range_start or date(annee, m, 1) > range_end:
                        continue
                    lo = 1 if m != range_start.month else max(1, range_start.day)
                    hi = days_in if m != range_end.month else min(days_in, range_end.day)
                else:
                    lo = start_day; hi = end_day if end_day <= days_in else days_in

                frames.append(
                    plan_for_single_month(
                        df, annee, m, freq_global,
                        freq_rand=freq_rand, freq_min=freq_min, freq_max=freq_max,
                        trace_global_qty=quota_global if trace_global else None,
                        mode=mode, seuil_min_kg=seuil_min,
                        order_mode=order_mode, alpha_=alpha, use_history=True,
                        seed_=seed+m, avoid_weekend=avoid_weekend, min_gap_days=min_gap_days,
                        daily_cap=daily_cap, start_day=lo, end_day=hi, tol_pct=tol_pct
                    )
                )
            plan = pd.concat(frames, ignore_index=True) if len(frames) else pd.DataFrame([])
        else:
            m = month_num(mois_label)
            days_in = calendar.monthrange(annee, m)[1]
            lo = start_day; hi = end_day if end_day <= days_in else days_in
            if use_calendar and isinstance(range_start, date) and isinstance(range_end, date):
                lo = range_start.day; hi = min(days_in, range_end.day)

            plan = plan_for_single_month(
                df, annee, m, freq_global,
                freq_rand=freq_rand, freq_min=freq_min, freq_max=freq_max,
                trace_global_qty=quota_global if trace_global else None,
                mode=mode, seuil_min_kg=seuil_min,
                order_mode=order_mode, alpha_=alpha, use_history=True,
                seed_=seed, avoid_weekend=avoid_weekend, min_gap_days=min_gap_days,
                daily_cap=daily_cap, start_day=lo, end_day=hi, tol_pct=tol_pct
            )

        if plan.empty:
            st.warning("Aucun plan généré (vérifie capacité, historique et paramètres).")
        else:
            monthly_summary = (
                plan.groupby(["Code plantation","Code","Nom plantation","Coopérative","Année","Mois"], as_index=False)
                    .agg(**{
                        "Plan mensuel (kg)": ("Plan mensuel (kg)", "max"),
                        "Prévu total mois (kg)": ("Quantité prévue (kg)","sum")
                    })
            )

            st.success("Plan généré ✅")
            st.subheader("Aperçu du plan (ordre choisi)")
            st.dataframe(plan.head(200), use_container_width=True, height=420)

            st.subheader("Résumé mensuel")
            st.dataframe(monthly_summary.head(100), use_container_width=True)

            # Export Excel (Plan + Suivi + Résumé + Paramètres)
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="xlsxwriter", datetime_format="yyyy-mm-dd", date_format="yyyy-mm-dd") as writer:
                plan.to_excel(writer, sheet_name="Plan", index=False)

                tracking = plan.copy()
                tracking["Quantité livrée (kg)"] = ""
                tracking["Livré à date (kg)"] = 0.0
                tracking["Livré annuel (kg)"] = 0.0
                tracking["Reste annuel (kg)"] = 0.0
                tracking["Livré du mois (kg)"] = 0.0
                tracking["Reste du mois (kg)"] = 0.0
                tracking["Alerte annuel"] = ""
                cols = ["Code plantation","Code","Nom plantation","Coopérative","Année","Mois","N° évènement","Date planifiée",
                        "Rendement annuel (kg)","Plan mensuel (kg)","Quantité prévue (kg)","Alerte cap",
                        "Quantité livrée (kg)","Livré à date (kg)","Livré annuel (kg)","Reste annuel (kg)",
                        "Livré du mois (kg)","Reste du mois (kg)","Alerte annuel"]
                tracking = tracking[cols]
                tracking.to_excel(writer, sheet_name="Suivi", index=False)
                ws = writer.sheets["Suivi"]
                n = len(tracking)
                for i in range(2, n+2):
                    ws.write_formula(i-1, 13, f'=IF($A{i}="","", SUMIFS($M:$M,$A:$A,$A{i}))')  # Livré annuel
                    ws.write_formula(i-1, 14, f'=IF($A{i}="","", $I{i} - $N{i})')              # Reste annuel
                    ws.write_formula(i-1, 15, f'=IF($A{i}="","", SUMPRODUCT(($A:$A=$A{i})*(MONTH($H:$H)=MONTH($H{i}))*(YEAR($H:$H)=YEAR($H{i}))*$M:$M))')  # Livré du mois
                    ws.write_formula(i-1, 16, f'=IF($A{i}="","", $J{i} - $O{i})')              # Reste du mois
                    ws.write_formula(i-1, 17, f'=IF($L{i}<0,"EXCÈS","OK")')                    # Alerte annuel

                monthly_summary.to_excel(writer, sheet_name="Résumé mensuel", index=False)
                params = pd.DataFrame({
                    "Paramètre": [
                        "Année","Mois","Fréquence globale","Fréq. aléatoire","Plage fréq.",
                        "Target actif","Quantité à tracer (kg)","Mode répartition","Seuil min (kg)",
                        "Éviter week-ends","Min gap (jours)","Fenêtre via calendrier","Cap journalier (kg)",
                        "Tolérance dépassement (%)","Dirichlet α","Historique activé","Seed","Ordre",
                        "Nb exclusions"
                    ],
                    "Valeur": [
                        annee, mois_label, freq_global, "Oui" if freq_rand else "Non", f"{freq_min}-{freq_max}",
                        "Oui" if trace_global else "Non", quota_global if trace_global else "", mode, seuil_min,
                        "Oui" if avoid_weekend else "Non", min_gap_days, "Oui" if use_calendar else "Non", daily_cap,
                        tol_pct, alpha, "Oui", seed, order_mode, len(excl_codes)
                    ]
                })
                params.to_excel(writer, sheet_name="Paramètres", index=False)

            st.download_button("💾 Télécharger l'Excel (Plan + Suivi + Résumé + Paramètres)",
                               data=buffer.getvalue(),
                               file_name=f"plan_livraisons_{annee}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            if auto_book:
                count = 0
                for _, rr in plan.iterrows():
                    if insert_delivery(rr["Code plantation"], rr["Nom plantation"], rr["Coopérative"], int(rr["Année"]), int(rr["Mois"]), rr["Date planifiée"], float(rr["Quantité prévue (kg)"])):
                        count += 1
                st.success(f"{count} lignes enregistrées comme 'livrées' dans l'historique.")

    # -------------------- Historique --------------------
    st.header("Historique des livraisons (base de données)")
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Ajouter une livraison (saisie rapide)")
        if df is not None and 'CODE_PLANTATION' in df.columns:
            # menu affiche CODE + NOM + COOP
            df_sel = df.copy()
            df_sel["__label__"] = df_sel["CODE_PLANTATION"].astype(str) + " | " + df_sel["CODE"].astype(str) + " | " + df_sel["NOM"].astype(str)
            label = st.selectbox("Choisir la plantation", df_sel["__label__"].tolist())
            row = df_sel[df_sel["__label__"]==label].iloc[0]
            code_sel = row["CODE_PLANTATION"]; nom_sel = row["NOM"]; coop_sel = row["COOPÉRATIVE"]
        else:
            code_sel = st.text_input("Code plantation (manuel)")
            nom_sel  = st.text_input("Nom plantation (manuel)")
            coop_sel = st.text_input("Coopérative (manuel)", value=coop_input)
        annee_in = st.number_input("Année livraison", min_value=2020, max_value=2035, value=int(annee))
        mois_in = st.selectbox("Mois livraison", ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"], index=0)
        mois_num_in = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"].index(mois_in)+1
        jour = st.number_input("Jour livraison", min_value=1, max_value=31, value=1)
        qty_in = st.number_input("Quantité livrée (kg)", min_value=0.0, value=0.0, step=10.0)
        if st.button("➕ Enregistrer la livraison"):
            try:
                dte = date(int(annee_in), int(mois_num_in), int(jour))
                if insert_delivery(code_sel, nom_sel, coop_sel, annee_in, mois_num_in, dte, qty_in):
                    st.success("Livraison enregistrée.")
                else:
                    st.info("Cette livraison existe déjà (même code/date/quantité).")
            except Exception as e:
                st.error(f"Erreur enregistrement : {e}")

    with col2:
        st.subheader("Importer un Excel 'Suivi'")
        upl_hist = st.file_uploader("Sélectionner un 'Suivi' exporté depuis l'outil", type=["xlsx"], key="import_hist")
        if upl_hist is not None:
            try:
                tdf = pd.read_excel(upl_hist, sheet_name="Suivi")
                needed = ["Code plantation","Nom plantation","Coopérative","Année","Mois","Date planifiée","Quantité livrée (kg)"]
                missing2 = [c for c in needed if c not in tdf.columns]
                if missing2:
                    st.error(f"Colonnes manquantes dans 'Suivi' : {missing2}")
                else:
                    tdf = tdf[pd.to_numeric(tdf["Quantité livrée (kg)"], errors="coerce").fillna(0) > 0]
                    cnt = 0
                    for _, rr in tdf.iterrows():
                        if insert_delivery(rr["Code plantation"], rr["Nom plantation"], rr["Coopérative"], int(rr["Année"]), int(rr["Mois"]), rr["Date planifiée"], float(rr["Quantité livrée (kg)"])):
                            cnt += 1
                    st.success(f"{cnt} lignes importées dans l'historique.")
            except Exception as e:
                st.error(f"Erreur import : {e}")

    st.subheader("Gestion avancée de l'historique")
    hcol1, hcol2, hcol3 = st.columns(3)
    with hcol1:
        if st.button("📤 Exporter l'historique (CSV)"):
            hist = fetch_history(limit=1000000)
            csv = hist.to_csv(index=False).encode("utf-8")
            st.download_button("Télécharger CSV", data=csv, file_name="historique_livraisons.csv", mime="text/csv", key="dl_csv")
    with hcol2:
        if st.button("📤 Exporter l'historique (Excel)"):
            hist = fetch_history(limit=1000000)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
                hist.to_excel(w, sheet_name="Historique", index=False)
            st.download_button("Télécharger Excel", data=buf.getvalue(), file_name="historique_livraisons.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xlsx")
    with hcol3:
        confirm = st.text_input("Écrire SUPPRIMER pour vider l'historique")
        if st.button("🗑️ Vider l'historique"):
            if confirm.strip().upper() == "SUPPRIMER":
                clear_history()
                st.success("Historique vidé.")
            else:
                st.warning("Confirmation manquante : tape 'SUPPRIMER'.")

# ==================== 10) FICHE D'EXPORT ====================
st.header("Fiche d'export — remplir le modèle à partir du plan")
with st.expander("Paramètres de la fiche d'export", expanded=True):
    colA, colB, colC = st.columns(3)
    with colA:
        exportateur_nom = st.text_input("Nom de l'exportateur", value="Bahari S.A.R.L.")
        export_lot_no = st.text_input("Export Lot N°", value="LOT-2025-001")
    with colB:
        connaissement = st.text_input("Connaissement (B/L)", value="BL-123456")
        destination = st.text_input("Destination (port/pays)", value="Abidjan, CI")
    with colC:
        contact = st.text_input("Contact / Référence", value="Opérations Export")
        commentaire = st.text_area("Commentaire (facultatif)", value="", height=95)
    tpl_file = st.file_uploader("Modèle Excel à compléter (facultatif)", type=["xlsx"], key="export_tpl")

def build_export_workbook(plan_df, meta, template_file=None):
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    import io
    if template_file is not None:
        try:
            wb = load_workbook(template_file)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()
    if "EXPORT_META" in wb.sheetnames: wb.remove(wb["EXPORT_META"])
    ws_meta = wb.create_sheet("EXPORT_META")
    ws_meta.append(["Clé","Valeur"])
    for k,v in [
        ("Exportateur", meta.get("exportateur_nom","")),
        ("Export Lot N°", meta.get("export_lot_no","")),
        ("Connaissement", meta.get("connaissement","")),
        ("Destination", meta.get("destination","")),
        ("Contact", meta.get("contact","")),
        ("Commentaire", meta.get("commentaire","")),
        ("Généré le", pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
    ]:
        ws_meta.append([k,v])

    if "PLANNING_EXPORT" in wb.sheetnames: wb.remove(wb["PLANNING_EXPORT"])
    ws_det = wb.create_sheet("PLANNING_EXPORT")
    cols = ["Export Lot N°","Connaissement","Exportateur","Année","Mois","N° évènement","Date planifiée",
            "Code plantation","Code","Nom plantation","Coopérative","Quantité prévue (kg)","Plan mensuel (kg)","Rendement annuel (kg)"]
    ws_det.append(cols)
    if len(plan_df) > 0:
        for _, r in plan_df.iterrows():
            ws_det.append([
                meta.get("export_lot_no",""), meta.get("connaissement",""), meta.get("exportateur_nom",""),
                int(r.get("Année","") or 0), int(r.get("Mois","") or 0), int(r.get("N° évènement",1) or 1),
                str(pd.to_datetime(r.get("Date planifiée")).date()) if r.get("Date planifiée","")!="" else "",
                str(r.get("Code plantation","")), str(r.get("Code","")), str(r.get("Nom plantation","")),
                str(r.get("Coopérative","")), float(r.get("Quantité prévue (kg)",0.0)), float(r.get("Plan mensuel (kg)",0.0)),
                float(r.get("Rendement annuel (kg)",0.0)),
            ])
    if "RESUME_JOUR" in wb.sheetnames: wb.remove(wb["RESUME_JOUR"])
    ws_sum = wb.create_sheet("RESUME_JOUR")
    if len(plan_df) > 0:
        pivot = (plan_df.groupby("Date planifiée", as_index=False)["Quantité prévue (kg)"].sum().sort_values("Date planifiée"))
        ws_sum.append(["Date","Total prévu (kg)"])
        for _, rr in pivot.iterrows():
            ws_sum.append([str(pd.to_datetime(rr["Date planifiée"]).date()), float(rr["Quantité prévue (kg)"])])

    for ws in [ws_meta, ws_det, ws_sum]:
        for col in range(1, ws.max_column+1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col)].width = 18

    out = io.BytesIO(); wb.save(out); return out.getvalue()

if 'plan' in locals() and isinstance(plan, pd.DataFrame) and not plan.empty:
    meta = {"exportateur_nom": exportateur_nom, "export_lot_no": export_lot_no, "connaissement": connaissement,
            "destination": destination, "contact": contact, "commentaire": commentaire}
    if st.button("📄 Construire la fiche d'export à partir du plan"):
        file_bytes = build_export_workbook(plan, meta, tpl_file)
        st.download_button("💾 Télécharger la fiche d'export (.xlsx)", data=file_bytes,
                           file_name=f"FICHE_EXPORT_{export_lot_no or 'LOT'}_{annee}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("👉 Génère d'abord un plan pour activer la fiche d'export.")
